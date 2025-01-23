import asyncio
import json
import os
import random
import sys
import traceback
from datetime import datetime, timedelta
from decimal import Decimal

from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Side, Border, Font, Alignment, PatternFill
from openpyxl.workbook import Workbook
from patchright._impl._errors import TargetClosedError
from patchright.async_api import BrowserContext, async_playwright

import settings
from models import Profile, Result
from settings import PROFILES_PATH, RESULTS_PATH, USER_FILES_FOLDER, ROOT_DIR

def get_usernames(txt_path):
    with open(txt_path) as f:
        usernames = f.readlines()
        return usernames

def get_accounts_from_excel(excel_path: str) -> list:
    profiles = []
    workbook = load_workbook(excel_path)

    sheet = workbook['not_done']
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=10, values_only=True):
        if not row[2]:
            continue
        # print(row)
        # If you put just auth token it will create cookie automatically
        if row[4] and len(row[4]) == 40:
            cookie = {'name': 'auth_token',
                      'value': row[4],
                      'domain': '.twitter.com', 'path': '/',
                      'expires': int((datetime.now() + timedelta(days=random.randint(300, 350))).timestamp()),
                      'httpOnly': True, 'secure': True, 'sameSite': 'None'}
        # or if you put cookie in a string it is loaded into json
        elif row[4] and len(row[4]) > 40:
            cookie = json.loads(row[4].replace("'", '"')
                                .replace('True', "true")
                                .replace('False', "false")
                                )
        else:
            cookie = None

        profile = Profile(name=row[0], proxy=row[1], seed=row[2], ref_code=row[3], cookie=cookie, x_username=row[5])

        profiles.append(profile)

    logger.info(f"Получил из таблицы profiles листа not_done {len(profiles)} профилей")

    workbook.close()
    return profiles


def write_results_for_profile(excel_path: str, profile: Profile, result: Result):
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # Создаем стиль для границ (все стороны)
    thin = Side(border_style="thin", color="000000")  # Тонкая черная линия
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Определяем номер строки для записи (следующая свободная строка)
    row_num = sheet.max_row + 1

    # Записываем данные в новую строку
    sheet.cell(row=row_num, column=1, value=profile.name).border = border
    sheet.cell(row=row_num, column=2, value=profile.proxy.as_url).border = border
    sheet.cell(row=row_num, column=3, value=profile.seed).border = border
    sheet.cell(row=row_num, column=4, value=profile.ref_code).border = border
    sheet.cell(row=row_num, column=5, value=json.dumps(profile.cookie)).border = border

    sheet.cell(row=row_num, column=6, value=result.bubble_amount).border = border
    sheet.cell(row=row_num, column=7, value=result.tasks_done).border = border
    sheet.cell(row=row_num, column=8, value=result.total_win_amount).border = border
    sheet.cell(row=row_num, column=9, value=datetime.now()).border = border

    workbook.save(excel_path)
    workbook.close()


def move_profile_to_done(excel_path: str, profile: Profile):
    workbook = load_workbook(excel_path)
    sheet_not_done = workbook['not_done']
    sheet_done = workbook['done']

    # Удаляем профиль из 'not_done' в обратном порядке
    rows = list(sheet_not_done.iter_rows(min_row=2, max_row=sheet_not_done.max_row, min_col=1, max_col=9, values_only=True))
    for i in range(len(rows), 0, -1):
        if rows[i-1][2] == profile.seed:  # row[2] - seed
            sheet_not_done.delete_rows(i + 1)  # +1, так как iter_rows начинается с 2-й строки

    # Удаляем пустые строки
    rows_done = list(sheet_done.iter_rows(min_row=2, max_row=sheet_done.max_row, min_col=1, max_col=9, values_only=True))
    for i in range(len(rows_done), 0, -1):
        if rows[i-1][1] is None:  # row[2] - seed
            sheet_done.delete_rows(i + 1)

    # Проверяем, есть ли профиль в 'done'
    rows_done = list(sheet_done.iter_rows(min_row=2, max_row=sheet_done.max_row, min_col=1, max_col=9, values_only=True))
    for row in rows_done:
        if row[2] == profile.seed:  # Профиль уже перенесен
            workbook.save(excel_path)
            workbook.close()
            return

    # Добавляем профиль в 'done'
    row = [profile.name, profile.proxy.as_url, profile.seed, profile.ref_code, json.dumps(profile.cookie)]
    sheet_done.append(row)

    workbook.save(excel_path)
    workbook.close()


def line_control(file_txt):
    # Удаление пустых строк
    with open(file_txt) as f1:
        lines = f1.readlines()
        non_empty_lines = (line for line in lines if not line.isspace())
        with open(file_txt, "w") as n_f1:
            n_f1.writelines(non_empty_lines)


def randfloat(from_: int | float | str, to_: int | float | str,
              step: int | float | str | None = None) -> float:
    """
    Return a random float from the range.

    :param Union[int, float, str] from_: the minimum value
    :param Union[int, float, str] to_: the maximum value
    :param Optional[Union[int, float, str]] step: the step size (calculated based on the number of decimal places)
    :return float: the random float
    """
    from_ = Decimal(str(from_))
    to_ = Decimal(str(to_))
    if not step:
        step = 1 / 10 ** (min(from_.as_tuple().exponent, to_.as_tuple().exponent) * -1)

    step = Decimal(str(step))
    rand_int = Decimal(str(random.randint(0, int((to_ - from_) / step))))
    return float(rand_int * step + from_)

def print_stats(stats: list[Result]):
    for result in stats:
        print(result)


def join_path(path: str | tuple | list) -> str:
    if isinstance(path, str):
        return path
    return str(os.path.join(*path))


def touch(path: str | tuple | list, file: bool = False) -> bool:
    """
    Create an object (file or directory) if it doesn't exist.

    :param Union[str, tuple, list] path: path to the object
    :param bool file: is it a file?
    :return bool: True if the object was created
    """
    path = join_path(path)
    if file:
        if not os.path.exists(path):
            with open(path, 'w') as f:
                f.write('')

            return True

        return False

    if not os.path.isdir(path):
        os.mkdir(path)
        return True

    return False


def create_files():
    touch(USER_FILES_FOLDER)
    touch(USER_FILES_FOLDER + '\profile_browsers\\')

    # Стиль заголовков
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="FFFF00")  # Желтый фон fgColor="FFFF00"
    header_fill_done = PatternFill("solid", fgColor="FF00FF")  # Желтый фон fgColor="FFFF00"

    # Создаем стиль для границ (все стороны)
    thin = Side(border_style="thin", color="000000")  # Тонкая черная линия
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if not os.path.exists(PROFILES_PATH):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "not_done"

        headers = ['name', 'proxy', 'seed', 'ref_code', 'cookie_or_token']

        # Записываем заголовки
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = border

        sheet_done = workbook.copy_worksheet(sheet)
        sheet_done.title = "done"
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill_done

        workbook.save(PROFILES_PATH)
        workbook.close()

    if not os.path.exists(RESULTS_PATH):
        workbook = Workbook()
        sheet = workbook.active

        headers = ['name', 'proxy', 'seed', 'ref_code', 'cookie_or_token',
                   'bubble_amount', 'tasks_done', 'total_win_amount', 'Time']

        # Записываем заголовки
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = border

        workbook.save(RESULTS_PATH)
        workbook.close()


create_files()

logger.remove()
logger.add(
    sys.stdout,
    # colorize=True,
    # format="<light-cyan>{time:HH:mm:ss}</light-cyan> | <level>{level: <8}</level> | <fg #ffffff>{name}:{line}</fg #ffffff> - <bold>{message}</bold>",
)
logger.add(
    os.path.join(ROOT_DIR, 'log.log'),
    # colorize=True,
    # format="<light-cyan>{time:HH:mm:ss}</light-cyan> | <level>{level: <8}</level> | <fg #ffffff>{name}:{line}</fg #ffffff> - <bold>{message}</bold>",
)
